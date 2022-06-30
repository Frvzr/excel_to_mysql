import mysql.connector
from mysql.connector import errorcode
import openpyxl
import time
from settings import user, password, path

start = time.time()
conn = mysql.connector.connect(user=user,  password=password)
cursor = conn.cursor()

DB_NAME = 'price'
TABLES = {}
TABLES['price'] = (
    "CREATE TABLE `price` (`id` int(20) NOT NULL AUTO_INCREMENT,"
    "`description` TEXT,"
    "`qty` int(11),"
    "`part_number` varchar(25),"
    "`articul` varchar(25),"
    "`price` decimal(10,2),"
    "`ea` varchar(14),"
    "PRIMARY KEY (`id`)"
") ENGINE=InnoDB")

def create_database(cursor):
    try:
        cursor.execute(
            "CREATE DATABASE {} DEFAULT CHARACTER SET 'utf8'".format(DB_NAME))
    except mysql.connector.Error as err:
        print("Failed creating database: {}".format(err))
        exit(1)

try:
    cursor.execute("USE {}".format(DB_NAME))
except mysql.connector.Error as err:
    print("Database {} does not exists.".format(DB_NAME))
    if err.errno == errorcode.ER_BAD_DB_ERROR:
        create_database(cursor)
        print("Database {} created successfully.".format(DB_NAME))
        conn.database = DB_NAME
    else:
        print(err)
        exit(1)

for table_name in TABLES:
    table_description = TABLES[table_name]
    try:
        print("Creating table {}: ".format(table_name), end='')
        cursor.execute(table_description)
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_TABLE_EXISTS_ERROR:
            print("already exists.")
        else:
            print(err.msg)
    else:
        print("OK")

wb = openpyxl.load_workbook(path)
sheet = wb.active
nb_row = sheet.max_row

# data = [[
# sheet[f'B{i}'].value,
# sheet[f'C{i}'].value,
# sheet[f'D{i}'].value,
# sheet[f'E{i}'].value,
# sheet[f'F{i}'].value,
# sheet[f'G{i}'].value
# ]
# for i in range (2, nb_row)
# ]

data = list(sheet.iter_rows(min_row=2, max_row=nb_row, max_col=6, values_only=True))
cursor.executemany("insert into `price`(description, qty, part_number, articul, price, ea) values(%s, %s, %s, %s, %s, %s)", data)

# for i in range (2, nb_row):
#     description = sheet[f'B{i}'].value
#     qty = sheet[f'C{i}'].value
#     part_number = sheet[f'D{i}'].value
#     articul = sheet[f'E{i}'].value
#     price = sheet[f'F{i}'].value
#     ea = sheet[f'G{i}'].value
#     cursor.execute("insert into `price`(description, qty, part_number, articul, price, ea) values(%s, %s, %s, %s, %s, %s)", (description, qty, part_number, articul, price, ea))

conn.commit()
cursor.close()
conn.close()
end = time.time()
print(end-start)